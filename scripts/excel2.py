# Importujemy niezbędne klasy do stylizacji z biblioteki openpyxl.
from openpyxl import Workbook
from openpyxl.styles import Font, Color, Alignment, PatternFill, Border, Side

# Tworzymy nowy skoroszyt (Workbook).
wb = Workbook()
# Wybieramy domyślnie utworzony arkusz.
ws = wb.active
# Nadajemy mu nazwę.
ws.title = "Formatowanie"

# 1. Zmiana czcionki (Font).
# Ustawiamy tekst w komórce A1, a następnie aplikujemy pogrubioną, czerwoną czcionkę o rozmiarze 16.
ws['A1'] = "Pogrubiony Czerwony"
ws['A1'].font = Font(name='Arial', size=16, bold=True, color="FF0000")

# 2. Wypełnienie komórki (Fill).
# Ustawiamy tło komórki B1 na kolor żółty.
ws['B1'] = "Żółte tło"
ws['B1'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# 3. Wyrównanie tekstu (Alignment).
# Komórka C1 będzie miała tekst wyśrodkowany zarówno w pionie, jak i w poziomie.
ws['C1'] = "Wyśrodkowany"
ws['C1'].alignment = Alignment(horizontal="center", vertical="center")

# 4. Obramowanie (Border).
# Definiujemy styl linii (cienka).
thin_side = Side(border_style="thin", color="000000")
# Nakładamy obramowanie na wszystkie krawędzie komórki D1.
ws['D1'] = "Obramowanie"
ws['D1'].border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)

# 5. Zmiana szerokości kolumny i wysokości wiersza.
# Ustawiamy szerokość kolumny C na 30 oraz wysokość pierwszego wiersza na 40.
ws.column_dimensions['C'].width = 30
ws.row_dimensions[1].height = 40

# Zapisujemy skoroszyt do pliku.
wb.save("scripts/Nazwisko_excel2.xlsx")
