# Importujemy klasę Workbook.
from openpyxl import Workbook

# Tworzymy nowy skoroszyt.
wb = Workbook()
ws = wb.active
ws.title = "Formuły i Scalanie"

# 1. Scalanie komórek (Merging cells).
# Łączymy komórki od A1 do D1.
ws.merge_cells('A1:D1')
ws['A1'] = "Nagłówek zestawienia danych"
# Możemy też rozłączyć komórki za pomocą ws.unmerge_cells('A1:D1').

# 2. Wprowadzanie danych testowych do obliczeń.
ws['A3'] = "Produkt"
ws['B3'] = "Cena"
ws['C3'] = "Ilość"
ws['D3'] = "Suma"

dane = [
    ("Jabłka", 3.50, 10),
    ("Gruszki", 4.20, 5),
    ("Banany", 2.80, 8)
]

# Wypełniamy wiersze danymi zaczynając od 4 wiersza.
for row_idx, (nazwa, cena, ilosc) in enumerate(dane, start=4):
    ws.cell(row=row_idx, column=1, value=nazwa)
    ws.cell(row=row_idx, column=2, value=cena)
    ws.cell(row=row_idx, column=3, value=ilosc)
    # 3. Używanie formuł Excela.
    # Ustawiamy formułę mnożącą cenę przez ilość (np. =B4*C4).
    ws.cell(row=row_idx, column=4, value=f"=B{row_idx}*C{row_idx}")

# 4. Suma całkowita na dole.
last_row = 4 + len(dane)
ws.cell(row=last_row, column=3, value="SUMA:")
# Formuła sumująca kolumnę D (np. =SUM(D4:D6)).
ws.cell(row=last_row, column=4, value=f"=SUM(D4:D{last_row-1})")

# Zapisujemy skoroszyt.
wb.save("scripts/Nazwisko_excel3.xlsx")
