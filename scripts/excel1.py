# Importujemy klasę Workbook z biblioteki openpyxl, która służy do tworzenia i edycji plików Excel.
from openpyxl import Workbook

# Tworzymy nową instancję skoroszytu (pliku Excel).
wb = Workbook()

# Tworzymy nowe arkusze o określonych nazwach.
ws1 = wb.create_sheet("Linki")
ws2 = wb.create_sheet("Filmweb")
ws3 = wb.create_sheet("Todos")

# Usuwamy domyślnie tworzony arkusz o nazwie 'Sheet'.
wb.remove(wb['Sheet'])

# Ustawiamy arkusz ws1 jako aktywny (ten, który otworzy się jako pierwszy po otwarciu pliku).
wb.active = ws1

# Wpisujemy wartość numeryczną do komórki A4 w arkuszu "Linki".
ws1['A4'] = 5

# Wpisujemy tekst do komórki A4 w arkuszu "Filmweb".
ws2['A4'] = "Hello"

# Zapisujemy skoroszyt do pliku o podanej nazwie.
wb.save("scripts/Nazwisko_excel1.xlsx")
