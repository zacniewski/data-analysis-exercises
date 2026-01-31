# Importujemy funkcję load_workbook do otwierania istniejących plików.
from openpyxl import load_workbook
import os

# Scieżka do pliku, który chcemy przeczytać.
# Użyjemy pliku wygenerowanego przez excel3.py, jeśli istnieje.
file_path = "scripts/Nazwisko_excel3.xlsx"

# Sprawdzamy, czy plik istnieje, aby uniknąć błędu.
if not os.path.exists(file_path):
    print(f"Błąd: Plik {file_path} nie istnieje. Uruchom najpierw excel3.py.")
else:
    # 1. Ładowanie istniejącego skoroszytu.
    # data_only=True sprawia, że w komórkach z formułami odczytamy wynik obliczeń, 
    # a nie samą treść formuły (wymaga to jednak, by plik był wcześniej zapisany przez Excela).
    wb = load_workbook(filename=file_path, data_only=False)
    ws = wb.active

    print(f"Czytanie arkusza: {ws.title}\n")

    # 2. Iteracja po wszystkich wierszach i kolumnach (ws.iter_rows).
    # min_row=3 zaczyna czytanie od nagłówków w 3 wierszu.
    print("Zawartość arkusza (wiersz po wierszu):")
    for row in ws.iter_rows(min_row=3, values_only=True):
        # 'row' jest krotką zawierającą wartości z poszczególnych komórek w wierszu.
        print(row)

    # 3. Odczyt konkretnej komórki.
    # Możemy odwoływać się przez adres (np. 'A4') lub współrzędne.
    komorka_produkt = ws['A4'].value
    komorka_cena = ws.cell(row=4, column=2).value
    
    print(f"\nDetal z wiersza 4: Produkt={komorka_produkt}, Cena={komorka_cena}")

    # 4. Sprawdzanie zakresu arkusza.
    print(f"\nZakres danych: {ws.dimensions}")
    print(f"Liczba wierszy: {ws.max_row}")
    print(f"Liczba kolumn: {ws.max_column}")

    # Zamykamy skoroszyt (dobre praktyki).
    wb.close()

# Dodatkowe zadanie: generowanie pliku tekstowego z Nazwiskiem
with open("scripts/Nazwisko_excel5.txt", "w") as f:
    f.write("Skrypt excel5.py został uruchomiony.")
